from decimal import Decimal
from openpyxl import Workbook, load_workbook
import random
import os

class JSON_to_excel:
	def __init__(self) -> None:
		pass

	def find_text(self, number):
		dict_values = {
			0:'',
			1:'Un',
			2:'Dos',
			3:'Tres',
			4:'Cuatro',
			5:'Cinco',
			6:'Seis',
			7:'Siete',
			8:'Ocho',
			9:'Nueve',
			10:'Diez',
			11:'Once',
			12:'Doce',
			13:'Trece',
			14:'Catorce',
			15:'Quince',
			16:'Dieciséis',
			17:'Diecisiete',
			18:'Dieciocho',
			19:'Diecinueve',
			20:'Veinte',
			21:'Veintiún',
			22:'Veintidós',
			23:'Veintitrés',
			24:'Veinticuatro',
			25:'Veinticinco',
			26:'Veintiséis',
			27:'Veintisiete',
			28:'Veintiocho',
			29:'Veintinueve',
			100:'Cien'
		}
		DECENAS = ['Treinta', 'Cuarenta', 'Cincuenta', 'Sesenta', 'Setenta', 'Ochenta', 'Noventa','']
		CIENTOS = ['Ciento','Doscientos','Trescientos','Cuatrocientos','Quinientos','Seiscientos', 'Setecientos','Ochocientos','Novecientos','']
		aux = ""
		try:
			print(dict_values[int(number)])
			return dict_values[int(number)]
		except:
			print('No dict Value')
			if len(number) == 2:
				aux = DECENAS[int(number[0])-3]
				if int(number[1]) != 0:
					aux = aux + ' y ' + dict_values[int(number[1])]
					return aux
				else:
					return aux
			elif len(number) == 3:
				aux = CIENTOS[int(number[0]) - 1] 
				if int(number[1]) != 0:
					split_string = number[1] + number[2]
					return aux + " " + self.find_text(split_string)
				else:
					return aux + ' ' + dict_values[int(number[2])]

	def get_text(self, number):

		quetzales = number[0]
		centavos = number[1]
		aux = ''
		if int(quetzales) > 1:
			if len(quetzales)/3 <= 1:
				if int(centavos) == 0:
					return self.find_text(quetzales) + ' Quetzales Exactos'
				else:
					return self.find_text(quetzales) + ' Quetzales con ' + self.find_text(centavos) + ' Centavos'
			elif len(quetzales)/3 <= 2:
				if len(quetzales) == 4:
					#print(quetzales[0], '+++', quetzales[1:] )
					aux = self.find_text(quetzales[0]) + " Mil " + self.find_text(quetzales[1:]) + " Quetzales" 
				elif len(quetzales) == 5:
					#print(quetzales[0:2], '+++', quetzales[2:])
					aux = self.find_text(quetzales[0:2]) + " Mil " + self.find_text(quetzales[2:]) + " Quetzales"
				elif len(quetzales) == 6:
					#print(quetzales[0:3],'+++', quetzales[3:])
					aux = self.find_text(quetzales[0:3]) + " Mil " + self.find_text(quetzales[3:]) + " Quetzales"	
				if int(centavos) == 0:
					return aux + " Exactos"
				else:
					return aux + " con " +self.find_text(centavos) + " Centavos"
			else:
				pass
		elif int(quetzales) < 1:
			pass
		else:
			return "Un Quetzal"

	def create_xlsx(self, nit, fecha, cliente, direccion, productos):
		total = 0
		wb = load_workbook('template.xlsx')
		ws = wb.worksheets[0]
		n = random.randint(1,400)

		try:
			print('adding to C13 name')
			ws['C13'] = cliente
			print('adding to C14 nit')
			ws['F14'] = 'NIT: ' + str(nit)
			print('adding to C15 adress')
			ws['C15'] = direccion
			try:
				fecha_array = fecha.split('-')
				ws['E12'].value = int(fecha_array[2])
				ws['F12'].value = int(fecha_array[1])
				ws['G12'].value = int(fecha_array[0])
			except:
				print('No valid date format')
			
			for i in range(len(productos)):
				fila = 17 + i	
				BCELL = 'B'+str(fila)
				CCELL = 'C'+str(fila)
				FCELL = 'F'+str(fila)

				FVALUE = Decimal(productos[i][2])
				ws[BCELL].value = int(productos[i][0])
				ws[CCELL].value = productos[i][1]
				ws[FCELL].value = FVALUE

				total += Decimal(int(productos[i][0])*FVALUE)
			print(total)
			try:
				ws['B43'] = self.get_text(str(total).split('.'))
			except:
				print('Error during numbers to text function')
			output_name = 'output_{}.xlsx'.format(n)
			wb.save(output_name)
			os.startfile(output_name)
		except:
			print('Doesnt pass')